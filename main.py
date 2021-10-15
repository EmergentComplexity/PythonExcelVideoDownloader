# YouTube Liked Video Downloader Bot
# uses pyTube,and OpenPyXl to download a playlist from youtube and store video titles and URLs in an .xlsx Excel spreadsheet
#
# The bot is able to look through the Excel spreadsheet before it downloads a video. By doing this the program can be ran multiple times and
# download only the new videos added to the playlist and update the spreadsheet to include the videos just downloaded
#
# Author: Peter Sanchez 
# Date Created: 8/28/2020
# Current Working Version: 1.00
# Version Date: 8/29/2020  
#---------------------------------------------------------------------------------------------------------------------------------------------------


import re                                                        # re- Regular Expressions module
import pytube                                                    # import PyTube to allow downloading a playlist
from pytube import Playlist                                      
from openpyxl import load_workbook                               # import OpenPyXl to allow reading and editing Excel spreadsheets
wb = load_workbook('C:\\Users\\Peter\\OneDrive\\Desktop\\Liked_Videos.xlsx')                                                          # path of the Excel spreadsheet with video data

ws = wb.active                                                   # ws is the active worksheet because it is the only worksheet in the  workbook wb

DOWNLOAD_DIR = 'C:\\Users\\Peter\\Downloads'                     # path to where the videos will be downloaded

playlist = Playlist('https://youtube.com/playlist?list=PLI78Xi-Ejw4FXbhQ-YRBubOtMhHAezPVS')       # URL of playlist


##consider removing this part## 
# this fixes the empty playlist.videos list
playlist._video_regex = re.compile(r"\"url\":\"(/watch\?v=[\w-]*)")

# start on the second row of the spreadsheet because the headings "Video Name:" and "Video Link" are on the first row of each respective column
initial_title = 'A2'                                            
initial_url = 'B2'
count = 0

# 'titlex' is the integer representation of the second character of 'initial_title'. when initial_title = 'A2', titlex = 2
# 'titlex' is an integer value of the current row in the spreadsheet
titlex = int(initial_title[1:])                                 
#print(titlex)



print(len(playlist.video_urls))                                # print the URLS of the videos in the playlist
downloadedvids = []                                            # list which will contain the titles of the videos that have already been downloaded in a previous execution of the code
for url in playlist.video_urls:
    print(url)
    yt = pytube.YouTube(url)
    
    roww = 1
    for Row in ws.rows:
        tittle = yt.title
        if ws.cell(row = roww, column=1).value == tittle:      # if the title of the video is already in the worksheet, append title to 'downloadedvids' list
            downloadedvids.append(ws.cell(row = roww, column=1).value)
            titlex = titlex + 1                                # increment titlex
        roww = roww +1                                         # go to the next row
        
    
    # if the video is not already in the spreadsheet, download the video and add it to the spreadsheet
    if yt.title not in downloadedvids :
        # Add video title and URL to the next free space in the spreadsheet
        txt_tit = "A{}"                                            # titles are located on column A
        txt_url = "B{}"                                            # URLs are located on column B
        title_pos = txt_tit.format(titlex)                         # Title will be added to column A, Row = titlex 
        url_pos = txt_url.format(titlex)                           # URL will be added to column B, Row = titlex
        #print(title_pos)
        #print(url_pos)
        ws[title_pos] = yt.title                                                                                    # copy the title of the video to its allocated cell
        ws[url_pos] = url                                                                                           # copy the URL of the video to its allocated cell
        try:    
            stream = yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution') .desc() .first()  # find the stream of the video to be downloaded
            stream.download(output_path=DOWNLOAD_DIR)                  # download the video to the specified directory
            wb.save('C:\\Users\\Peter\\OneDrive\\Desktop\\Liked_Videos.xlsx')  # save the excel workbook
            print("Downloaded: ", url)                                 # print to confirm video download
            titlex = titlex + 1                                        # increment titlex
        except:
            print('error in downloading: ', url)                       # some error in downloading, Video was Deleted ?

  
